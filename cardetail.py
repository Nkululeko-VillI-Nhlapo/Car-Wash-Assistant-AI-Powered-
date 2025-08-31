import os
import json
from flask import Flask, request, jsonify
import anthropic
from twilio.rest import Client
import openpyxl
from datetime import datetime, timedelta, timezone
import logging
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
import io
import tempfile
import re
import random

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class IntelligentBusinessAssistant:
    def __init__(self):
        # Configuration
        self.anthropic_api_key = os.getenv('ANTHROPIC_API_KEY', #################")
        self.twilio_account_sid = os.getenv('TWILIO_ACCOUNT_SID', "#################")
        self.twilio_auth_token = os.getenv('TWILIO_AUTH_TOKEN', "#################")
        self.twilio_whatsapp_number = "whatsapp:+14155238886"

       
        # Google Drive configuration
        self.google_drive_file_id = "17Uj56WrAf-FXfyFHhZMk7efyDYyMNWOr"
        self.credentials_file = "cardetail-ai-959d1fe801f2.json"
        
        # Initialize services
        self.anthropic_client = anthropic.Anthropic(api_key=self.anthropic_api_key)
        self.twilio_client = Client(self.twilio_account_sid, self.twilio_auth_token)
        self.drive_service = self.init_google_drive()
        
        # Enhanced conversation memory with context
        self.conversations = {}
        
        # Service pricing (for smart suggestions)
        self.service_prices = {
            'Basic Wash': 120,
            'Full Wash': 180,
            'Premium Wash': 280,
            'Interior Only': 150
        }
        
    def init_google_drive(self):
        """Initialize Google Drive service"""
        try:
            scopes = ['https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(self.credentials_file, scopes=scopes)
            return build('drive', 'v3', credentials=creds)
        except Exception as e:
            logger.error(f"Error initializing Google Drive: {str(e)}")
            return None
    
    def download_excel_from_drive(self):
        """Download Excel file from Google Drive"""
        try:
            request = self.drive_service.files().get_media(fileId=self.google_drive_file_id)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            downloader = MediaIoBaseDownload(temp_file, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
            temp_file.close()
            return temp_file.name
        except Exception as e:
            logger.error(f"Error downloading from Google Drive: {str(e)}")
            return None
    
    def upload_excel_to_drive(self, local_file_path):
        """Upload Excel file to Google Drive"""
        try:
            media = MediaFileUpload(local_file_path, 
                                  mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            file = self.drive_service.files().update(
                fileId=self.google_drive_file_id, media_body=media).execute()
            return True
        except Exception as e:
            logger.error(f"Error uploading to Google Drive: {str(e)}")
            return False
    
    def get_sa_datetime(self):
        """Get current South African date and time"""
        sa_timezone = timezone(timedelta(hours=2))  # South Africa is UTC+2
        return datetime.now(sa_timezone)
    
    def calculate_week_of_month(self, date):
        """Calculate week number within the month (1-4)"""
        first_day_of_month = date.replace(day=1)
        days_from_first = (date - first_day_of_month).days
        return (days_from_first // 7) + 1
    
    def load_business_data(self):
        """Load comprehensive business data with updated column structure"""
        try:
            local_file = self.download_excel_from_drive()
            if not local_file:
                return None
            
            workbook = openpyxl.load_workbook(local_file)
            
            # Load Operations data with updated structure
            operations_data = []
            operations_sheet = workbook['Operations']
            for row in operations_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Customer_ID exists
                    operations_data.append({
                        'Customer_ID': row[0], 'Customer_Name': row[1] or '',
                        'Service_Date': row[2], 'Week_Number': row[3] or 1,
                        'Month': row[4] or datetime.now().strftime('%B'),
                        'Service_Completed': row[5] or 'No', 'Service_Type': row[6] or '',
                        'Notes': row[7] or '', 'Status': row[8] or 'Pending'
                    })
            
            # Load Revenue data with new column structure
            revenue_data = []
            revenue_sheet = workbook['Revenue']
            for row in revenue_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    revenue_data.append({
                        'Transaction_Id': row[0], 'Customer_id': row[1], 'Service_Date': row[2],
                        'Month': row[3], 'Service_Type': row[4], 'Amount': row[5] or 0,
                        'Payment_Status': row[6] or 'Unpaid', 'Payment_Method': row[7] or 'Cash',
                        'Status': row[8] or 'Pending', 'Week_Number': row[9] or 1
                    })
            
            # Load Expenses data (structure unchanged)
            expenses_data = []
            expenses_sheet = workbook['Expenses']
            for row in expenses_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    expenses_data.append({
                        'Transaction_ID': row[0], 'Date': row[1], 'Month': row[2],
                        'Category': row[3] or 'Other', 'Description': row[4] or '',
                        'Amount': abs(row[5]) if row[5] else 0, 'Supplier': row[6] or '',
                        'Status': row[7] or 'Pending', 'Notes': row[8] or ''
                    })
            
            # Advanced cross-sheet analysis
            total_revenue = sum(item['Amount'] for item in revenue_data)
            total_expenses = sum(item['Amount'] for item in expenses_data)
            net_profit = total_revenue - total_expenses
            profit_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0
            
            # Service performance analysis
            service_performance = {}
            for item in revenue_data:
                service = item['Service_Type']
                if service not in service_performance:
                    service_performance[service] = {'revenue': 0, 'count': 0, 'customers': set()}
                service_performance[service]['revenue'] += item['Amount']
                service_performance[service]['count'] += 1
                service_performance[service]['customers'].add(item['Customer_id'])
            
            # Convert sets to counts for JSON serialization
            for service in service_performance:
                service_performance[service]['unique_customers'] = len(service_performance[service]['customers'])
                service_performance[service]['customers'] = list(service_performance[service]['customers'])
            
            # Customer journey analysis
            customer_journey = {}
            for op in operations_data:
                customer_id = op['Customer_ID']
                if customer_id not in customer_journey:
                    customer_journey[customer_id] = {
                        'name': op['Customer_Name'],
                        'services_completed': 0,
                        'total_spent': 0,
                        'last_service_date': None,
                        'service_types': [],
                        'payment_methods': [],
                        'notes': []
                    }
                
                customer_journey[customer_id]['services_completed'] += 1
                customer_journey[customer_id]['service_types'].append(op['Service_Type'])
                customer_journey[customer_id]['notes'].append(op['Notes'])
                customer_journey[customer_id]['last_service_date'] = op['Service_Date']
                
                # Match with revenue data
                for rev in revenue_data:
                    if rev['Customer_id'] == customer_id:
                        customer_journey[customer_id]['total_spent'] += rev['Amount']
                        customer_journey[customer_id]['payment_methods'].append(rev['Payment_Method'])
            
            # Expense analysis
            expense_analysis = {}
            for item in expenses_data:
                category = item['Category']
                expense_analysis[category] = expense_analysis.get(category, 0) + item['Amount']
            
            # Service completion and payment status
            completed_services = len([op for op in operations_data if op['Service_Completed'] == 'Yes'])
            completion_rate = (completed_services / len(operations_data) * 100) if operations_data else 0
            
            paid_revenue = sum(item['Amount'] for item in revenue_data if item['Payment_Status'] == 'Paid')
            payment_rate = (paid_revenue / total_revenue * 100) if total_revenue > 0 else 0
            
            os.unlink(local_file)
            
            return {
                'operations_data': operations_data, 'revenue_data': revenue_data, 'expenses_data': expenses_data,
                'total_revenue': total_revenue, 'total_expenses': total_expenses,
                'net_profit': net_profit, 'profit_margin': profit_margin,
                'service_performance': service_performance, 'customer_journey': customer_journey,
                'expense_analysis': expense_analysis, 'completion_rate': completion_rate,
                'payment_rate': payment_rate, 'paid_revenue': paid_revenue,
                'total_customers': len(customer_journey),
                'repeat_customers': len([c for c in customer_journey.values() if c['services_completed'] > 1])
            }
            
        except Exception as e:
            logger.error(f"Error loading business data: {str(e)}")
            return None
    
    def generate_customer_id(self):
        """Generate next customer ID"""
        try:
            data = self.load_business_data()
            if data and data['operations_data']:
                max_num = 0
                for op in data['operations_data']:
                    if op['Customer_ID'] and op['Customer_ID'].startswith('CW'):
                        try:
                            num = int(op['Customer_ID'][2:])
                            max_num = max(max_num, num)
                        except:
                            continue
                return f"CW{str(max_num + 1).zfill(3)}"
            else:
                return "CW001"
        except:
            return f"CW{str(random.randint(100, 999))}"
    
    def extract_service_info(self, message):
        """Extract comprehensive service information with intelligent prompting"""
        # Service type detection
        services = {
            'Basic Wash': ['basic', 'simple', 'quick wash', 'standard', 'normal'],
            'Full Wash': ['full', 'complete', 'comprehensive', 'thorough'],
            'Premium Wash': ['premium', 'deluxe', 'luxury', 'detailed', 'top service'],
            'Interior Only': ['interior', 'inside', 'cabin clean', 'inside only']
        }
        
        service_type = None
        message_lower = message.lower()
        for service, keywords in services.items():
            if any(keyword in message_lower for keyword in keywords):
                service_type = service
                break
        
        # Payment method detection
        payment_methods = {
            'Cash': ['cash', 'notes', 'money'],
            'Card': ['card', 'swipe', 'tap'],
            'EFT': ['eft', 'transfer', 'bank transfer', 'online']
        }
        
        payment_method = 'Cash'  # default
        for method, keywords in payment_methods.items():
            if any(keyword in message_lower for keyword in keywords):
                payment_method = method
                break
        
        # Extract customer name
        customer_patterns = [
            r'(?:customer|client|for|served)\s+([A-Za-z]+(?:\s+[A-Za-z]+)?)',
            r'([A-Za-z]+(?:\s+[A-Za-z]+)?)(?:\s+came|paid|wants)',
            r'(?:mr|mrs|ms)\.?\s+([A-Za-z]+)',
        ]
        
        customer_name = None
        for pattern in customer_patterns:
            match = re.search(pattern, message, re.IGNORECASE)
            if match:
                potential_name = match.group(1).strip().title()
                excluded_words = ['Service', 'Wash', 'Customer', 'Full', 'Basic', 'Premium', 'Interior', 'Cash', 'Card']
                if potential_name not in excluded_words:
                    customer_name = potential_name
                    break
        
        # Extract amount
        amount = None
        amount_patterns = [
            r'[rR]\s*(\d+(?:,\d+)*(?:\.\d{2})?)',
            r'(\d+(?:,\d+)*(?:\.\d{2})?)\s*rand',
            r'\b(\d+(?:,\d+)*(?:\.\d{2})?)\b',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, message)
            if matches:
                try:
                    amounts = [float(match.replace(',', '')) for match in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        # Smart amount suggestion
        if not amount and service_type and service_type in self.service_prices:
            amount = self.service_prices[service_type]
        
        # Payment status detection
        payment_status = 'Paid'  # default for completed services
        if any(word in message_lower for word in ['unpaid', 'owes', 'later', 'credit']):
            payment_status = 'Unpaid'
        
        return {
            'customer_name': customer_name,
            'service_type': service_type,
            'amount': amount,
            'payment_method': payment_method,
            'payment_status': payment_status
        }
    
    def extract_expense_info(self, message):
        """Intelligently extract expense information"""
        # Extract amounts
        amount_patterns = [
            r'[rR]\s*(\d+(?:,\d+)*(?:\.\d{2})?)',
            r'(\d+(?:,\d+)*(?:\.\d{2})?)\s*rand',
            r'\b(\d+(?:,\d+)*(?:\.\d{2})?)\b',
        ]
        
        amount = None
        for pattern in amount_patterns:
            matches = re.findall(pattern, message)
            if matches:
                try:
                    amounts = [float(match.replace(',', '')) for match in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        # Extract supplier
        supplier_patterns = [
            r'(?:paid|pay|from|to|bought from|purchased from)\s+([A-Za-z][A-Za-z\s]+?)(?:\s|$|for|,|\d)',
            r'([A-Za-z][A-Za-z\s]+(?:suppliers?|store|shop|company|ltd|pty|co))',
            r'(?:supplier|vendor):\s*([A-Za-z][A-Za-z\s]+)',
        ]
        
        supplier = None
        for pattern in supplier_patterns:
            match = re.search(pattern, message, re.IGNORECASE)
            if match:
                potential_supplier = match.group(1).strip()
                potential_supplier = re.sub(r'\s+', ' ', potential_supplier)
                if len(potential_supplier) > 1 and not potential_supplier.isdigit():
                    supplier = potential_supplier
                    break
        
        # Category detection
        category_keywords = {
            'Supplies': ['soap', 'chemical', 'detergent', 'wax', 'towel', 'bucket', 'supplies', 'cleaning', 'brushes'],
            'Equipment': ['hose', 'machine', 'vacuum', 'pressure', 'equipment', 'tool', 'replacement'],
            'Utilities': ['water', 'electricity', 'power', 'utility', 'bill'],
            'Staff': ['salary', 'wage', 'pay', 'employee', 'worker', 'staff'],
            'Marketing': ['advert', 'marketing', 'promotion', 'flyer', 'social media'],
            'Fuel': ['petrol', 'diesel', 'fuel', 'gas'],
            'Maintenance': ['repair', 'fix', 'service', 'maintenance', 'upkeep']
        }
        
        category = 'Other'
        message_lower = message.lower()
        for cat, keywords in category_keywords.items():
            if any(keyword in message_lower for keyword in keywords):
                category = cat
                break
        
        return {
            'amount': amount,
            'supplier': supplier,
            'category': category,
            'description': message[:100].strip()
        }
    
    def ask_for_missing_info(self, data_type, extracted_info, phone_number):
        """Ask for missing information in column order"""
        if data_type == 'service':
            missing_info = []
            
            if not extracted_info.get('customer_name'):
                missing_info.append("What's the customer's name?")
            
            if not extracted_info.get('service_type'):
                missing_info.append("What type of service? (Basic Wash, Full Wash, Premium Wash, or Interior Only)")
                
            if not extracted_info.get('amount'):
                missing_info.append("What's the service amount in Rand?")
            
            if missing_info:
                return f"I need a few more details, Moloi: {' '.join(missing_info)}"
            
        elif data_type == 'expense':
            missing_info = []
            
            if not extracted_info.get('amount'):
                missing_info.append("How much was the expense in Rand?")
                
            if not extracted_info.get('supplier'):
                missing_info.append("Who did you pay?")
                
            if not extracted_info.get('category') or extracted_info.get('category') == 'Other':
                missing_info.append("What category? (Supplies, Equipment, Utilities, Staff, Marketing, Fuel, or Maintenance)")
            
            if missing_info:
                return f"I need more details for the expense, Moloi: {' '.join(missing_info)}"
        
        return None
    
    def save_complete_service(self, service_data):
        """Save service to both Operations and Revenue sheets with updated structure"""
        try:
            local_file = self.download_excel_from_drive()
            if not local_file:
                return None
                
            workbook = openpyxl.load_workbook(local_file)
            
            # Generate customer ID
            customer_id = self.generate_customer_id()
            
            # Get SA date and time
            sa_now = self.get_sa_datetime()
            current_date = sa_now.date()
            current_week = self.calculate_week_of_month(current_date)
            current_month = sa_now.strftime('%B')
            
            # Save to Operations sheet (9 columns)
            operations_sheet = workbook['Operations']
            ops_row = operations_sheet.max_row + 1
            
            operations_sheet.cell(row=ops_row, column=1, value=customer_id)  # Customer_ID
            operations_sheet.cell(row=ops_row, column=2, value=service_data['customer_name'] or 'Walk-in Customer')  # Customer_Name
            operations_sheet.cell(row=ops_row, column=3, value=current_date)  # Service_Date
            operations_sheet.cell(row=ops_row, column=4, value=current_week)  # Week_Number
            operations_sheet.cell(row=ops_row, column=5, value=current_month)  # Month
            operations_sheet.cell(row=ops_row, column=6, value='Yes')  # Service_Completed
            operations_sheet.cell(row=ops_row, column=7, value=service_data['service_type'] or 'General Service')  # Service_Type
            operations_sheet.cell(row=ops_row, column=8, value='Logged via WhatsApp')  # Notes
            operations_sheet.cell(row=ops_row, column=9, value='Completed')  # Status
            
            # Save to Revenue sheet (10 columns with new structure)
            revenue_sheet = workbook['Revenue']
            rev_row = revenue_sheet.max_row + 1
            transaction_id = f"REV{str(rev_row-1).zfill(3)}"
            
            revenue_sheet.cell(row=rev_row, column=1, value=transaction_id)  # Transaction_Id
            revenue_sheet.cell(row=rev_row, column=2, value=customer_id)  # Customer_id
            revenue_sheet.cell(row=rev_row, column=3, value=current_date)  # Service_Date
            revenue_sheet.cell(row=rev_row, column=4, value=current_month)  # Month
            revenue_sheet.cell(row=rev_row, column=5, value=service_data['service_type'] or 'General Service')  # Service_Type
            revenue_sheet.cell(row=rev_row, column=6, value=float(service_data['amount']))  # Amount
            revenue_sheet.cell(row=rev_row, column=7, value=service_data.get('payment_status', 'Paid'))  # Payment Status
            revenue_sheet.cell(row=rev_row, column=8, value=service_data['payment_method'])  # Payment_Method
            revenue_sheet.cell(row=rev_row, column=9, value='Washed' if service_data.get('payment_status') == 'Paid' else 'Not yet Washed')  # Status
            revenue_sheet.cell(row=rev_row, column=10, value=current_week)  # Week_Number (Column 10)
            
            workbook.save(local_file)
            upload_success = self.upload_excel_to_drive(local_file)
            os.unlink(local_file)
            
            return {
                'customer_id': customer_id,
                'transaction_id': transaction_id,
                'success': upload_success
            }
        except Exception as e:
            logger.error(f"Error saving complete service: {str(e)}")
            return None
    
    def save_expense(self, expense_data):
        """Save expense with updated structure"""
        try:
            local_file = self.download_excel_from_drive()
            if not local_file:
                return None
                
            workbook = openpyxl.load_workbook(local_file)
            sheet = workbook['Expenses']
            
            next_row = sheet.max_row + 1
            transaction_id = f"EXP{str(next_row-1).zfill(3)}"
            
            # Get SA date
            sa_now = self.get_sa_datetime()
            current_date = sa_now.date()
            current_month = sa_now.strftime('%B')
            
            sheet.cell(row=next_row, column=1, value=transaction_id)  # Transaction_ID
            sheet.cell(row=next_row, column=2, value=current_date)  # Date
            sheet.cell(row=next_row, column=3, value=current_month)  # Month
            sheet.cell(row=next_row, column=4, value=expense_data['category'])  # Category
            sheet.cell(row=next_row, column=5, value=expense_data['description'])  # Description
            sheet.cell(row=next_row, column=6, value=-abs(float(expense_data['amount'])))  # Amount (negative)
            sheet.cell(row=next_row, column=7, value=expense_data['supplier'] or 'Unknown')  # Supplier/Receiver
            sheet.cell(row=next_row, column=8, value='Recorded')  # Status
            sheet.cell(row=next_row, column=9, value='Added via WhatsApp')  # Notes
            
            workbook.save(local_file)
            upload_success = self.upload_excel_to_drive(local_file)
            os.unlink(local_file)
            
            return transaction_id if upload_success else None
        except Exception as e:
            logger.error(f"Error saving expense: {str(e)}")
            return None
    
    def analyze_cash_flow(self):
        """Simple cash flow analysis"""
        data = self.load_business_data()
        if not data:
            return "Can't access business data right now, Moloi."
        
        cash_in = data['total_revenue']
        cash_out = data['total_expenses']
        net_cash_flow = cash_in - cash_out
        
        if net_cash_flow > 0:
            return f"Cash flow is positive, Moloi. R{cash_in:,.0f} coming in, R{cash_out:,.0f} going out. Net cash flow: R{net_cash_flow:,.0f}."
        else:
            return f"Cash flow is negative, Moloi. R{cash_in:,.0f} coming in, R{cash_out:,.0f} going out. You're short R{abs(net_cash_flow):,.0f}."
    
    def create_simple_income_statement(self):
        """Basic income statement"""
        data = self.load_business_data()
        if not data:
            return "Can't access business data, Moloi."
        
        revenue = data['total_revenue']
        expenses = data['total_expenses'] 
        net_income = revenue - expenses
        
        report = f"INCOME STATEMENT:\n"
        report += f"Revenue: R{revenue:,.0f}\n"
        report += f"Expenses: R{expenses:,.0f}\n"
        report += f"Net Income: R{net_income:,.0f}\n"
        
        if net_income > 0:
            report += f"You made a profit of R{net_income:,.0f}"
        else:
            report += f"You made a loss of R{abs(net_income):,.0f}"
            
        return report
    
    def explain_finance_term(self, term):
        """Explain financial terms"""
        explanations = {
            'profit': "Money left after paying all expenses. Revenue minus expenses.",
            'loss': "When you spend more than you earn. Expenses are higher than revenue.",
            'revenue': "All money coming into the business from customers.",
            'income': "Same as revenue - money earned from services.",
            'expenses': "All money spent to run the business - supplies, utilities, wages.",
            'cash flow': "Money moving in and out of business. Positive = more in than out.",
            'margin': "Profit as percentage of revenue. Shows how profitable each rand of sales is.",
            'break even': "Point where revenue equals expenses. No profit, no loss.",
        }
        
        term_lower = term.lower()
        if term_lower in explanations:
            return f"{term.title()}: {explanations[term_lower]}"
        else:
            return f"Not sure about that term, Moloi. Ask me about profit, revenue, expenses, cash flow, or margins."
    
    def generate_enhanced_report(self):
        """Generate business insights with new metrics"""
        data = self.load_business_data()
        if not data:
            return "Sorry Moloi, can't access business data right now. Try again in a moment."
        
        # Financial summary
        if data['net_profit'] > 0:
            report = f"Business is profitable, Moloi. Revenue R{data['total_revenue']:,.0f}, expenses R{data['total_expenses']:,.0f}. You made R{data['net_profit']:,.0f} profit."
        else:
            report = f"Business is making a loss, Moloi. Revenue R{data['total_revenue']:,.0f}, expenses R{data['total_expenses']:,.0f}. Loss of R{abs(data['net_profit']):,.0f}."
        
        # Payment status
        report += f" Payment rate: {data['payment_rate']:.1f}% (R{data['paid_revenue']:,.0f} collected)."
        
        # Service performance
        if data['service_performance']:
            best_service = max(data['service_performance'].items(), key=lambda x: x[1]['revenue'])
            report += f" Your best service is {best_service[0]} with R{best_service[1]['revenue']:,.0f} from {best_service[1]['count']} jobs."
        
        # Customer insight
        report += f" You served {data['total_customers']} customers, {data['repeat_customers']} came back for more."
        
        # Recommendations
        if data['payment_rate'] < 90:
            report += " Focus on collecting payments - some customers still owe money."
        elif data['profit_margin'] < 10:
            report += " Profit is low - consider increasing prices or cutting expenses."
        
        return report
    
    def process_natural_message(self, message, phone_number):
        """Enhanced message processing with updated column awareness"""
        context = self.conversations.get(phone_number, {'history': []})
        context['history'].append({'user': message, 'timestamp': datetime.now()})
        
        business_data = self.load_business_data()
        
        # Get SA date for context
        sa_now = self.get_sa_datetime()
        today_sa = sa_now.strftime('%d %B %Y')
        
        business_context = ""
        if business_data:
            business_context = f"""
REAL-TIME BUSINESS STATUS:
Financial: R{business_data['total_revenue']:,.0f} revenue | R{business_data['total_expenses']:,.0f} expenses | R{business_data['net_profit']:,.0f} profit
Payment Status: {business_data['payment_rate']:.1f}% payments collected (R{business_data['paid_revenue']:,.0f})
Operations: {len(business_data['operations_data'])} services | {business_data['completion_rate']:.1f}% completion rate
"""
        
        system_prompt = f"""You are Moloi's business assistant for MR Banks Car Detailing in South Africa.

CONTEXT AWARENESS:
- Today's date in South Africa: {today_sa}
- Currency: South African Rand (R)
- Business location: South Africa
- Week numbers: 1-4 within each month

COMMUNICATION STYLE:
- Always call him "Moloi"
- Keep responses SHORT and practical 
- Use simple, clear English
- Be direct and helpful
- Use "Rand" or "R" for currency

SMART DATA COLLECTION:
- Services: Customer Name, Service Type, Amount (Rand), Payment Method, Payment Status
- Expenses: Amount (Rand), Supplier, Category
- Extract from full sentences when possible
- Auto-generate IDs and calculate dates/weeks
- Track payment status (Paid/Unpaid)

CURRENT BUSINESS DATA:
{business_context}

DETECTION RULES:
- Service: served, customer, wash, service, finished, completed
- Expense: paid, bought, spent, cost, expense, bill, salary
- NEVER log purchases as services
- Track both service completion AND payment status

Be conversational but focused on accurate data collection."""
        
        try:
            response = self.anthropic_client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=200,
                system=system_prompt,
                messages=[{"role": "user", "content": message}]
            )
            
            ai_response = response.content[0].text.strip()
            
            message_lower = message.lower()
            
            # Expense detection
            expense_indicators = ['paid', 'bought', 'spent', 'cost', 'expense', 'bill', 'salary']
            service_indicators = ['served', 'customer', 'wash', 'service', 'finished', 'completed']
            
            if (any(indicator in message_lower for indicator in expense_indicators) and
                not any(indicator in message_lower for indicator in service_indicators) and
                any(char.isdigit() for char in message)):
                
                expense_info = self.extract_expense_info(message)
                missing_info_msg = self.ask_for_missing_info('expense', expense_info, phone_number)
                
                if missing_info_msg:
                    ai_response = missing_info_msg
                elif expense_info['amount']:
                    transaction_id = self.save_expense(expense_info)
                    if transaction_id:
                        ai_response = f"Expense logged: {transaction_id} - R{expense_info['amount']:,.0f} to {expense_info['supplier']} for {expense_info['category']}."
            
            # Service detection
            elif (any(indicator in message_lower for indicator in service_indicators) and
                  any(char.isdigit() for char in message) and
                  ('customer' in message_lower or 'client' in message_lower)):
                
                service_info = self.extract_service_info(message)
                missing_info_msg = self.ask_for_missing_info('service', service_info, phone_number)
                
                if missing_info_msg:
                    ai_response = missing_info_msg
                elif service_info['amount']:
                    service_result = self.save_complete_service(service_info)
                    if service_result and service_result['success']:
                        ai_response = f"Service logged: {service_result['customer_id']} ({service_info['customer_name']}) - {service_info['service_type']} R{service_info['amount']:,.0f} ({service_info['payment_status']})."
            
            # Business reporting
            elif any(trigger in message_lower for trigger in ['business', 'report', 'numbers', 'performance']):
                ai_response = self.generate_enhanced_report()
            
            # Financial analysis
            elif 'cash flow' in message_lower:
                ai_response = self.analyze_cash_flow()
            elif any(term in message_lower for term in ['income statement', 'profit and loss']):
                ai_response = self.create_simple_income_statement()
            elif any(term in message_lower for term in ['what is', 'explain']) and any(finance_term in message_lower for finance_term in ['profit', 'loss', 'revenue', 'margin']):
                finance_terms = ['profit', 'loss', 'revenue', 'income', 'expenses', 'cash flow', 'margin']
                for term in finance_terms:
                    if term in message_lower:
                        ai_response = self.explain_finance_term(term)
                        break
            
            context['history'].append({'assistant': ai_response, 'timestamp': datetime.now()})
            self.conversations[phone_number] = context
            
            return ai_response
            
        except Exception as e:
            logger.error(f"Error in Anthropic processing: {str(e)}")
            return "Sorry Moloi, I'm having a technical moment. Could you repeat that?"
    
    def send_whatsapp_message(self, to_number, message):
        """Send WhatsApp message"""
        try:
            sent_message = self.twilio_client.messages.create(
                body=message, 
                from_=self.twilio_whatsapp_number, 
                to=to_number
            )
            logger.info(f"Message sent to {to_number}")
            return True
        except Exception as e:
            logger.error(f"Error sending WhatsApp: {str(e)}")
            return False

# Flask app
app = Flask(__name__)
assistant = IntelligentBusinessAssistant()

@app.route('/webhook', methods=['POST'])
def webhook():
    """Handle incoming WhatsApp messages"""
    try:
        incoming_msg = request.values.get('Body', '').strip()
        from_number = request.values.get('From', '')
        
        logger.info(f"Received from {from_number}: {incoming_msg}")
        
        response = assistant.process_natural_message(incoming_msg, from_number)
        assistant.send_whatsapp_message(from_number, response)
        
        return jsonify({'status': 'success'}), 200
        
    except Exception as e:
        logger.error(f"Webhook error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/test', methods=['GET'])
def test():
    """Test endpoint"""
    return jsonify({
        'status': 'Enhanced Car Detailing Assistant with Updated Column Structure is running!',
        'features': [
            'Operations Tracking with Payment Status', 
            'Revenue Logging with Payment Tracking', 
            'Expense Management', 
            'Customer Journey Analysis', 
            'South African Business Context',
            'Smart Data Collection'
        ],
        'timestamp': datetime.now().isoformat()
    })

if __name__ == '__main__':
    print("Starting Enhanced MR Banks Car Detailing Assistant...")
    print("NEW FEATURES:")
    print("   - Updated Column Structure Support")
    print("   - Payment Status Tracking")
    print("   - South African Date/Time Context") 
    print("   - Smart Week Number Calculation (1-4 per month)")
    print("   - Enhanced Business Intelligence")
    app.run(host='0.0.0.0', port=5000, debug=True)